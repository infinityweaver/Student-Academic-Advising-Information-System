# -*- coding: utf-8 -*-
"""
One-time generator for per-student advising Markdown files.

Reads:  raw/<student-no>.json        (registrar grade scrapes)
        reference/curriculum/*.xlsx  (BSCS curricula, 2018-2024 and 2025-present)
        students/active/*/*.xlsx     (contact info from existing checklists)
Writes: students/active/<Folder>/<NAME>.md   and prints a roster summary.

WARNING: the MD files are meant to be hand-maintained after generation.
Re-running this script OVERWRITES them, including any advising notes.
"""
import json, os, re, glob, unicodedata, sys
from collections import defaultdict

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
GENERATED_ON = "2026-07-04"
LATEST_TERM = ("2025-2026", "Second Semester")  # most recent term in the scrapes

SEM_ORDER = {"First Semester": 1, "Second Semester": 2, "Summer": 3, "Mid Year": 3, "Midyear": 3}

def term_key(ay, sem):
    return (ay, SEM_ORDER.get(sem, 9))

def strip_accents(s):
    return "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")

def norm_code(code):
    return re.sub(r"\s+", " ", code.strip()).upper()

def base_code(code):
    # drop single trailing section-variant letter: "PhEd 11c" -> "PHED 11"
    return re.sub(r"(\d)[A-Z]$", r"\1", norm_code(code))

# ---------------------------------------------------------------- curricula
def load_curriculum(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Student"]
    courses, thresholds = [], []
    year, cur_term = 1, None
    mode = "courses"
    for row in ws.iter_rows(min_row=9, values_only=True):
        a, b, c = row[0], row[1], row[2]
        units = row[5]
        if mode == "courses":
            if c is None or str(c).strip() == "":
                if row[4] and "Current Total Units" in str(row[4]):
                    mode = "between"
                continue
            t = str(a).strip() if a else ""
            if t and t != cur_term:
                if t == "1st" and cur_term in ("2nd", "Midyear"):
                    year += 1
                cur_term = t
            courses.append({
                "code": str(b).strip(), "title": re.sub(r"\s+", " ", str(c)).strip(),
                "units": float(units) if units not in (None, "") else 0.0,
                "prereq": re.sub(r"\s+", " ", str(row[6])).strip() if row[6] else "None",
                "year": year, "term": cur_term,
            })
        else:
            if a is not None and str(a).strip().replace(".", "").isdigit() and b:
                thresholds.append((float(a), str(b).strip()))
    return {"courses": courses, "thresholds": sorted(thresholds)}

CUR = {
    "2018": load_curriculum(os.path.join(ROOT, "reference", "curriculum", "Course-Checklist-2018-2024.xlsx")),
    "2025": load_curriculum(os.path.join(ROOT, "reference", "curriculum", "Course-Checklist-2025-present.xlsx")),
}

def year_level(units, thresholds):
    label = thresholds[0][1]
    for u, lab in thresholds:
        if units >= u:
            label = lab
    return label

# ---------------------------------------------------------------- students
def load_contact(folder):
    xs = [f for f in glob.glob(os.path.join(folder, "*.xlsx")) if not os.path.basename(f).startswith("~")]
    if not xs:
        return {}
    try:
        wb = openpyxl.load_workbook(xs[0], data_only=True)
        ws = wb[wb.sheetnames[0]]
        return {
            "email": str(ws["G4"].value or "").strip(),
            "contact": str(ws["G6"].value or "").strip(),
        }
    except Exception as e:
        print("  ! contact read failed:", xs[0], e)
        return {}

def grade_kind(g):
    g = (g or "").strip()
    if g == "":
        return "none"
    if g == "INC":
        return "inc"
    if g in ("DR", "DRP"):
        return "dr"
    if g == "S":
        return "s"
    if g == "NA":
        return "na"
    try:
        float(g)
        return "num"
    except ValueError:
        return "other"

def effective(rec):
    """Effective outcome of one enrollment: (passed, numeric_grade_or_None, display)."""
    kind = grade_kind(rec["grade"])
    if kind == "inc":
        comp = (rec["completion"] or "").strip()
        if comp:
            try:
                v = float(comp)
                return (v <= 3.0, v, f"INC → {comp}")
            except ValueError:
                return (False, None, f"INC ({comp})")
        return (False, None, "INC (outstanding)")
    if kind == "num":
        v = float(rec["grade"])
        return (v <= 3.0, v, rec["grade"])
    if kind == "s":
        return (True, None, "S")
    if kind == "dr":
        return (False, None, "DR")
    if kind == "na":
        return (False, None, "NA")
    if kind == "none":
        return (False, None, "—")
    return (False, None, rec["grade"])

def short_term(ay, sem):
    s = {"First Semester": "1st Sem", "Second Semester": "2nd Sem", "Summer": "Summer",
         "Mid Year": "Midyear", "Midyear": "Midyear"}.get(sem, sem)
    return f"{s} {ay}"

def analyze(data):
    grades = data["grades"]
    sid = data["student"]["student_number"]
    curkey = "2025" if any(norm_code(g["course_code"]).startswith("CSIT") for g in grades) else "2018"
    cur = CUR[curkey]

    # group attempts per base code, chronological
    attempts = defaultdict(list)
    for g in sorted(grades, key=lambda r: term_key(r["academic_year"], r["semester"])):
        attempts[base_code(g["course_code"])].append(g)

    # match curriculum rows -> attempts (duplicate codes e.g. CSci 200 matched in order)
    used = defaultdict(int)  # base code -> how many attempts consumed by earlier duplicate rows
    dup_codes = defaultdict(list)
    for row in cur["courses"]:
        dup_codes[base_code(row["code"])].append(row)

    matched_recs = set()
    checklist = []
    for row in cur["courses"]:
        bc = base_code(row["code"])
        recs = attempts.get(bc, [])
        if len(dup_codes[bc]) > 1 and recs:
            # split attempts among duplicate rows by units when possible, else order
            recs = [r for r in recs if abs(r["units"] - row["units"]) < 0.01] or recs
            i = used[bc]
            recs = recs[i:i + 1]
            used[bc] += 1
        for r in recs:
            matched_recs.add(id(r))
        checklist.append({"row": row, "recs": recs})

    extra = [g for g in grades if id(g) not in matched_recs]

    # ---- aggregates ----
    passed_codes = set()
    units_earned = 0.0
    gwa_num = gwa_den = 0.0
    per_term = defaultdict(list)
    for g in grades:
        per_term[(g["academic_year"], g["semester"])].append(g)
        ok, val, _ = effective(g)
        if val is not None:
            gwa_num += val * g["units"]
            gwa_den += g["units"]
    for code, recs in attempts.items():
        best = None
        for r in recs:
            ok, val, _ = effective(r)
            if ok:
                best = r
        if best is not None:
            passed_codes.add(code)
            units_earned += best["units"]
    gwa = gwa_num / gwa_den if gwa_den else None

    # ---- flags ----
    flags = []
    # retakes: curriculum courses attempted, never passed, latest attempt final & failed
    for item in checklist:
        recs = item["recs"]
        if not recs:
            continue
        bc = base_code(item["row"]["code"])
        if bc in passed_codes:
            continue
        last = recs[-1]
        kind = grade_kind(last["grade"])
        if kind == "none":
            continue  # still in progress
        if kind == "inc" and not (last["completion"] or "").strip():
            flags.append(("INC", f"**{last['course_code']}** — INC from {short_term(last['academic_year'], last['semester'])};"
                                 f" must be completed within 1 year or it lapses to 5.00"))
        elif kind in ("num", "dr", "na"):
            _, val, disp = effective(last)
            flags.append(("RETAKE", f"**{last['course_code']}** {item['row']['title']} — {disp} in "
                                    f"{short_term(last['academic_year'], last['semester'])}; retake required"))
    # non-curriculum failures
    for g in extra:
        ok, val, disp = effective(g)
        kind = grade_kind(g["grade"])
        if kind in ("num", "dr", "na") and not ok and base_code(g["course_code"]) not in passed_codes:
            flags.append(("RETAKE", f"**{g['course_code']}** {g['course_title'].title()} — {disp} in "
                                    f"{short_term(g['academic_year'], g['semester'])}; not in checklist, verify if still required"))
        if kind == "inc" and not (g["completion"] or "").strip():
            flags.append(("INC", f"**{g['course_code']}** — INC from {short_term(g['academic_year'], g['semester'])};"
                                 f" must be completed within 1 year or it lapses to 5.00"))

    # delinquency per term (>=25% of enrolled units failed, final grades only)
    delinquent_terms = []
    for (ay, sem), recs in sorted(per_term.items(), key=lambda kv: term_key(*kv[0])):
        tot = sum(r["units"] for r in recs)
        failed = sum(r["units"] for r in recs
                     if grade_kind(r["grade"]) in ("num", "na") and not effective(r)[0]
                     and grade_kind(r["grade"]) != "none")
        if tot and failed / tot >= 0.25:
            delinquent_terms.append((ay, sem, failed, tot))
    for ay, sem, f_, t_ in delinquent_terms:
        flags.append(("DELINQ", f"Failed {f_:g}/{t_:g} units in {short_term(ay, sem)} (≥25% — delinquency rule)"))

    # enrollment currency
    last_term = max(per_term.keys(), key=lambda k: term_key(*k))
    enrolled_now = term_key(*last_term) >= term_key(*LATEST_TERM)
    if not enrolled_now:
        flags.append(("STOPOUT", f"No enrollment record since {short_term(*last_term)} — verify LOA/returnee status"))
    in_progress = [g for g in per_term.get(LATEST_TERM, []) if grade_kind(g["grade"]) == "none"]

    # status light
    recent_delinq = any(term_key(ay, sem) >= term_key("2025-2026", "First Semester") for ay, sem, _, _ in delinquent_terms)
    n_retakes = sum(1 for k, _ in flags if k == "RETAKE")
    if not enrolled_now or recent_delinq or n_retakes >= 3:
        status = "🔴 Needs attention"
    elif flags:
        status = "🟡 Watch"
    else:
        status = "🟢 On track"

    return {
        "curkey": curkey, "cur": cur, "checklist": checklist, "extra": extra,
        "attempts": attempts, "passed": passed_codes, "units": units_earned,
        "gwa": gwa, "flags": flags, "per_term": per_term, "status": status,
        "last_term": last_term, "in_progress": in_progress,
        "year_level": year_level(units_earned, cur["thresholds"]),
    }

# ---------------------------------------------------------------- rendering
STATUS_ICON = {"passed": "✅", "inc": "⏳", "failed": "❌", "progress": "✳️", "not": "⬜"}

def row_status(item, passed_codes):
    recs = item["recs"]
    if not recs:
        return ("⬜ Not taken", "", "")
    bc = base_code(item["row"]["code"])
    last = recs[-1]
    when = short_term(last["academic_year"], last["semester"])
    hist = "; ".join(f"{effective(r)[2]} ({short_term(r['academic_year'], r['semester'])})" for r in recs[:-1])
    ok, val, disp = effective(last)
    kind = grade_kind(last["grade"])
    if bc in passed_codes and ok:
        st = "✅ Passed"
    elif bc in passed_codes:
        st = "✅ Passed"  # passed on a different attempt/row with the same code
        ok2 = [r for r in recs if effective(r)[0]]
        if ok2:
            last = ok2[-1]; _, _, disp = effective(last)
            when = short_term(last["academic_year"], last["semester"])
    elif kind == "none":
        st = "✳️ In progress"
        disp = f"midterm {last['midterm']}" if last.get("midterm") else "—"
    elif kind == "inc" and not (last["completion"] or "").strip():
        st = "⏳ INC"
    else:
        st = "❌ Failed" if kind in ("num", "na") else "❌ Dropped"
    note = hist
    return (st, f"{disp} · {when}", note)

def render_md(data, info, an, folder_name):
    s = data["student"]
    sid = s["student_number"]
    name = s["name"].rstrip(" .")
    cur_label = "2018–2024 (old)" if an["curkey"] == "2018" else "2025–present (new)"
    entry_ay = f"20{sid[:2]}-20{int(sid[:2]) + 1}"
    email = info.get("email") or f"{sid}@vsu.edu.ph"
    contact = info.get("contact") or "—"
    grad_units = an["cur"]["thresholds"][-1][0]

    L = []
    L.append(f"# {name}")
    L.append("")
    L.append(f"> Status: **{an['status']}** · last updated {GENERATED_ON} (generated from registrar scrape `raw/{sid}.json`)")
    L.append("")
    L.append("| | |")
    L.append("|---|---|")
    L.append(f"| Student No. | {sid} |")
    L.append(f"| Program | {s['program']} |")
    L.append(f"| Curriculum | {cur_label} |")
    L.append(f"| Entered | AY {entry_ay} |")
    L.append(f"| Email | {email} |")
    L.append(f"| Contact | {contact} |")
    L.append(f"| Units earned | {an['units']:g} / {grad_units:g} |")
    L.append(f"| Year level (by units) | {an['year_level']} |")
    L.append(f"| GWA (all final numeric grades) | {an['gwa']:.2f} |" if an["gwa"] else "| GWA | — |")
    L.append(f"| Last term with records | {short_term(*an['last_term'])} |")
    L.append("")

    L.append("## Flags")
    L.append("")
    if an["flags"]:
        order = {"STOPOUT": 0, "DELINQ": 1, "INC": 2, "RETAKE": 3}
        for k, txt in sorted(an["flags"], key=lambda f: order.get(f[0], 9)):
            icon = {"RETAKE": "❌", "INC": "⏳", "DELINQ": "📉", "STOPOUT": "🚪"}[k]
            L.append(f"- {icon} {txt}")
    else:
        L.append("- None 🎉")
    L.append("")

    if an["in_progress"]:
        L.append(f"## Currently enrolled — no final grade yet ({short_term(*LATEST_TERM)})")
        L.append("")
        for g in an["in_progress"]:
            mt = f" (midterm {g['midterm']})" if g.get("midterm") else ""
            L.append(f"- {g['course_code']} — {g['course_title'].title()}{mt}")
        L.append("")

    L.append("## Curriculum checklist")
    L.append("")
    L.append("Grade rules: 1.00 best – 3.00 pass; 5.00 fail; INC must be completed within 1 year; DR = dropped.")
    L.append("")
    cur_year = None
    TERM_LABEL = {"1st": "1st Semester", "2nd": "2nd Semester", "Midyear": "Midyear"}
    cur_term = None
    for item in an["checklist"]:
        r = item["row"]
        if (r["year"], r["term"]) != (cur_year, cur_term):
            if cur_year is not None:
                L.append("")
            cur_year, cur_term = r["year"], r["term"]
            L.append(f"### Year {cur_year} — {TERM_LABEL.get(cur_term, cur_term)}")
            L.append("")
            L.append("| Course | Title | Units | Prerequisite | Status | Grade · When | Prior attempts |")
            L.append("|---|---|---|---|---|---|---|")
        st, gw, hist = row_status(item, an["passed"])
        L.append(f"| {r['code']} | {r['title']} | {r['units']:g} | {r['prereq']} | {st} | {gw} | {hist} |")
    L.append("")

    if an["extra"]:
        L.append("## Other courses taken (not matched to checklist)")
        L.append("")
        L.append("| Course | Title | Units | Grade | When |")
        L.append("|---|---|---|---|---|")
        for g in sorted(an["extra"], key=lambda r: term_key(r["academic_year"], r["semester"])):
            _, _, disp = effective(g)
            L.append(f"| {g['course_code']} | {g['course_title'].title()} | {g['units']:g} | {disp} | {short_term(g['academic_year'], g['semester'])} |")
        L.append("")

    L.append("## Grade history by term")
    L.append("")
    for (ay, sem), recs in sorted(an["per_term"].items(), key=lambda kv: term_key(*kv[0])):
        num = den = 0.0
        for g in recs:
            _, val, _ = effective(g)
            if val is not None:
                num += val * g["units"]; den += g["units"]
        twa = f" · term GWA {num / den:.2f}" if den else ""
        tot = sum(g["units"] for g in recs)
        L.append(f"### {short_term(ay, sem)} — {tot:g} units{twa}")
        L.append("")
        L.append("| Course | Title | Units | Midterm | Final | Completion |")
        L.append("|---|---|---|---|---|---|")
        for g in recs:
            L.append(f"| {g['course_code']} | {g['course_title'].title()} | {g['units']:g} |"
                     f" {g['midterm'] or '—'} | {g['grade'] or '—'} | {g['completion'] or '—'} |")
        L.append("")

    L.append("## Advising notes")
    L.append("")
    L.append("<!-- Hand-maintained. Add newest entries at the top. -->")
    L.append("")
    L.append("| Date | Note |")
    L.append("|---|---|")
    L.append(f"| {GENERATED_ON} | File generated from registrar scrape; verified against checklist workbook. |")
    L.append("")
    return "\n".join(L)

# ---------------------------------------------------------------- main
def main():
    active_dir = os.path.join(ROOT, "students", "active")
    folders = {strip_accents(f.split(",")[0]).upper(): f for f in os.listdir(active_dir)}
    roster = []
    for jf in sorted(glob.glob(os.path.join(ROOT, "raw", "*.json"))):
        data = json.load(open(jf, encoding="utf-8"))
        s = data["student"]
        surname = strip_accents(s["name"].split(",")[0]).upper()
        folder = folders.get(surname)
        if not folder:
            print("NO FOLDER for", s["name"]); sys.exit(1)
        fpath = os.path.join(active_dir, folder)
        info = load_contact(fpath)
        an = analyze(data)
        md = render_md(data, info, an, folder)
        out = os.path.join(fpath, folder.upper() + ".md")
        with open(out, "w", encoding="utf-8") as fh:
            fh.write(md)
        roster.append({
            "folder": folder, "id": s["student_number"], "name": s["name"].rstrip(" ."),
            "cur": an["curkey"], "units": an["units"], "gwa": an["gwa"],
            "year": an["year_level"], "status": an["status"],
            "flags": len(an["flags"]), "md": folder.upper() + ".md",
        })
        print(f"wrote {folder}/{folder.upper()}.md  [{an['status']}] units={an['units']:g} gwa={an['gwa'] and round(an['gwa'],2)}")
    with open(os.path.join(ROOT, "tools", "roster.json"), "w", encoding="utf-8") as fh:
        json.dump(roster, fh, indent=1, ensure_ascii=False)

if __name__ == "__main__":
    main()
