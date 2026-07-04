# -*- coding: utf-8 -*-
"""One-time migration: MD-as-database → record.json per student folder.

Sources, in priority order:
- students/<status>/<Folder>/<FOLDER>.md — profile header fields + advising notes
- raw/<student-no>.json                  — grade entries (registrar scrape)
- students/<status>/<Folder>/<NAME>.xlsx — email/contact fallback (contact sheet)

The MD files are left in place (they become an export artifact; regenerate
them any time via Export MD). Run:  python -m saais.migrate [--force]
"""
import argparse
import os
import re
import sys

from .repo import curriculum, md_doc, paths, records, scrape


def _find_md(folder_path, folder):
    conventional = os.path.join(folder_path, folder.upper() + ".md")
    if os.path.exists(conventional):
        return conventional
    mds = [f for f in os.listdir(folder_path) if f.lower().endswith(".md")]
    return os.path.join(folder_path, mds[0]) if mds else None


def _xlsx_profile(folder_path):
    """{name, sid, email, contact} from the per-student contact sheet, best effort."""
    out = {"name": "", "sid": "", "email": "", "contact": ""}
    xs = [f for f in os.listdir(folder_path) if f.lower().endswith(".xlsx") and not f.startswith("~$")]
    if not xs:
        return out
    try:
        import openpyxl
        wb = openpyxl.load_workbook(os.path.join(folder_path, xs[0]), data_only=True)
        ws = wb[wb.sheetnames[0]]
        cell = lambda r, c: str(ws.cell(row=r, column=c).value or "").strip()
        out["name"] = cell(4, 1)      # A4: Name
        out["sid"] = cell(6, 1)       # A6: Student No
        out["email"] = cell(4, 7)     # G4: E-mail
        out["contact"] = cell(6, 7)   # G6: Contact No
    except Exception:
        pass
    return out


def migrate_folder(status, folder, folder_path, force=False):
    """Returns (action, detail): action in created/overwritten/skipped/error."""
    existing = os.path.exists(paths.record_path(folder_path))
    if existing and not force:
        return "skipped", "record.json already exists"

    fields, notes, title_name = {}, [], None
    md = _find_md(folder_path, folder)
    if md:
        with open(md, encoding="utf-8") as fh:
            text = fh.read()
        doc = md_doc.parse(text)
        fields = md_doc.header_fields(doc["preamble"])
        notes = md_doc.parse_notes(text)   # newest first
        first = text.splitlines()[0] if text else ""
        if first.startswith("# "):
            title_name = first[2:].strip()

    xp = _xlsx_profile(folder_path)
    sid = fields.get("Student No.", "").strip() or xp["sid"]
    raw = None
    if sid:
        try:
            raw = scrape.load(sid)
        except scrape.ScrapeError as e:
            return "error", f"raw/{sid}.json invalid: {e}"
    if not sid:
        return "error", "no student number in MD or contact sheet — nothing to migrate"

    name = (raw["student"]["name"] if raw else title_name or xp["name"] or folder).rstrip(" .")
    program = raw["student"]["program"] if raw else fields.get("Program", "BSCS")
    email, contact = fields.get("Email", ""), fields.get("Contact", "")
    if not email or email == sid:
        email = xp["email"]
    contact = contact or xp["contact"]

    entered = re.sub(r"^AY\s+", "", fields.get("Entered", "")).strip()
    if not entered and sid[:2].isdigit():
        entered = f"20{sid[:2]}-20{int(sid[:2]) + 1}"

    rec = records.new_record(
        sid or folder, name, program,
        curriculum=curriculum.key_from_label(fields.get("Curriculum", "")),
        entered=entered, email=email, contact=contact, status=status)
    if raw:
        rec["grades"] = raw["grades"]
    if status == "graduated":
        year = os.path.basename(os.path.dirname(folder_path))
        rec["student"]["graduated_year"] = int(year) if year.isdigit() else None

    for i, (when, text) in enumerate(reversed(notes), start=1):  # oldest → newest
        rec["notes"].append({"id": i, "date": when, "text": text})
    rec["meta"]["note_seq"] = len(notes)

    records.save(folder_path, rec)   # backs up any existing record.json first
    n = len(rec["grades"])
    return ("overwritten" if existing else "created"), f"{n} grade entries, {len(notes)} notes"


LEGACY_CURRICULA = {  # v1 key -> (program, effective_start, effective_end)
    "2018": ("BSCS", 2018, 2024),
    "2025": ("BSCS", 2025, None),
}


def migrate_curricula(force=False):
    """Seed data/curricula/<key>.json from the legacy workbooks, keeping the
    v1 keys so existing student records keep resolving."""
    for key, xp in paths.CURRICULA.items():
        target = os.path.join(paths.CURRICULA_DIR, f"{key}.json")
        if os.path.exists(target) and not force:
            print(f"[    skipped] curriculum {key} — already migrated")
            continue
        if not os.path.exists(xp):
            print(f"[    skipped] curriculum {key} — {os.path.basename(xp)} not found")
            continue
        program, start, end = LEGACY_CURRICULA.get(key, ("BSCS", int(key), None))
        cur = curriculum.from_xlsx(xp, key, program, start, end)
        curriculum.save(cur)
        n = sum(len(s["courses"]) for s in cur["sections"])
        print(f"[    created] curriculum {key} — {curriculum.label(cur)}, {n} courses")


def main(argv=None):
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument("--force", action="store_true",
                    help="rewrite record.json even where it already exists (previous version is backed up)")
    args = ap.parse_args(argv)

    migrate_curricula(force=args.force)
    counts = {}
    for status, folder, fp in paths.student_dirs():
        try:
            action, detail = migrate_folder(status, folder, fp, force=args.force)
        except (records.RecordError, OSError, ValueError) as e:
            action, detail = "error", str(e)
        counts[action] = counts.get(action, 0) + 1
        print(f"[{action:>11}] {status}/{folder} — {detail}")
    print("\nSummary:", ", ".join(f"{v} {k}" for k, v in sorted(counts.items())) or "no student folders found")
    return 1 if counts.get("error") else 0


if __name__ == "__main__":
    sys.exit(main())
