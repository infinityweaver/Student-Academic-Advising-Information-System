# -*- coding: utf-8 -*-
"""First-class curriculum records — data/curricula/<id>.json.

Curricula contain no PII, so unlike student records they are committed. Each
record holds program, effective year range, and sections (year level +
semester) of courses. After creation only the effective year range is
editable — students' checklists depend on the sections and courses.

The legacy Course-Checklist-*.xlsx workbooks are importable (parse_xlsx), and
`python -m saais.migrate` seeds data/curricula/ from them under their v1 keys
("2018", "2025") so existing student records keep resolving.
"""
import json
import os
import re

from . import backups, paths

SCHEMA = 1
TERMS = ("1st", "2nd", "Midyear")

_cache = {"token": None, "data": None}


class CurriculumError(ValueError):
    """Raised when a curriculum does not match the expected schema."""


# ------------------------------------------------------------------- helpers
def label(cur):
    """Display title: '<program> <effective-year-range>', e.g. 'BSCS 2018–2024'."""
    end = cur["effective_end"]
    return f"{cur['program']} {cur['effective_start']}–{end if end else 'present'}"


def flatten(cur):
    """Course rows in the shape rules.analyze expects:
    [{code, title, units, prereq, year, term}] in section order."""
    rows = []
    for sec in cur["sections"]:
        for c in sec["courses"]:
            rows.append({**c, "year": sec["year"], "term": sec["term"]})
    return rows


def total_units(cur):
    return sum(c["units"] for sec in cur["sections"] for c in sec["courses"])


def default_thresholds(cur):
    """Quartile thresholds ending at (total units, 'Graduating') — used when a
    curriculum is created without explicit thresholds."""
    tot = total_units(cur)
    return [[0, "1st year"], [round(tot * .25), "2nd year"], [round(tot * .5), "3rd year"],
            [round(tot * .75), "4th year"], [tot, "Graduating"]]


def validate(cur):
    if not isinstance(cur, dict):
        raise CurriculumError("Curriculum must be an object.")
    for k in ("id", "program", "effective_start"):
        if not cur.get(k):
            raise CurriculumError(f"curriculum.{k} is missing or empty.")
    if not re.match(r"^[A-Za-z0-9_-]+$", str(cur["id"])):
        raise CurriculumError("Curriculum id must be a plain slug (letters/digits/-/_).")
    try:
        cur["effective_start"] = int(cur["effective_start"])
        cur["effective_end"] = int(cur["effective_end"]) if cur.get("effective_end") else None
    except (TypeError, ValueError):
        raise CurriculumError("Effective years must be numbers (end empty = present).")
    if cur["effective_end"] and cur["effective_end"] < cur["effective_start"]:
        raise CurriculumError("Effective end year is before the start year.")
    if not isinstance(cur.get("sections"), list) or not cur["sections"]:
        raise CurriculumError("A curriculum needs at least one section.")
    for i, sec in enumerate(cur["sections"]):
        try:
            sec["year"] = int(sec["year"])
        except (TypeError, ValueError):
            raise CurriculumError(f"sections[{i}].year must be a number.")
        if sec.get("term") not in TERMS:
            raise CurriculumError(f"sections[{i}].term must be one of {TERMS}.")
        if not isinstance(sec.get("courses"), list) or not sec["courses"]:
            raise CurriculumError(f"sections[{i}] ('Year {sec['year']} – {sec['term']}') has no courses.")
        for j, c in enumerate(sec["courses"]):
            for k in ("code", "title"):
                if not c.get(k):
                    raise CurriculumError(f"sections[{i}].courses[{j}].{k} is missing.")
            try:
                c["units"] = float(c["units"])
            except (TypeError, ValueError):
                raise CurriculumError(f"course {c.get('code')}: units is not a number.")
            c.setdefault("prereq", "None")
    if not cur.get("thresholds"):
        cur["thresholds"] = default_thresholds(cur)
    try:
        cur["thresholds"] = sorted([float(u), str(lab)] for u, lab in cur["thresholds"])
    except (TypeError, ValueError):
        raise CurriculumError("Thresholds must be (units, label) pairs.")
    return cur


# ------------------------------------------------------------------- storage
def _path(cid):
    return os.path.join(paths.CURRICULA_DIR, f"{cid}.json")


def _token():
    files = []
    if os.path.isdir(paths.CURRICULA_DIR):
        for f in sorted(os.listdir(paths.CURRICULA_DIR)):
            if f.endswith(".json"):
                fp = os.path.join(paths.CURRICULA_DIR, f)
                files.append((f, os.path.getmtime(fp)))
    return tuple(files)


def cache_token():
    """Changes whenever any curriculum file changes — used in index cache keys."""
    return _token()


def load_all():
    """{id: curriculum} with rules-compatible extras: 'courses' (flattened rows)
    and 'thresholds' as (units, label) tuples. Cached until a file changes."""
    token = _token()
    if _cache["token"] == token and _cache["data"] is not None:
        return _cache["data"]
    out = {}
    for fname, _ in token:
        fp = os.path.join(paths.CURRICULA_DIR, fname)
        with open(fp, encoding="utf-8") as fh:
            try:
                cur = validate(json.load(fh))
            except (json.JSONDecodeError, CurriculumError) as e:
                raise CurriculumError(f"{fname}: {e}")
        cur["courses"] = flatten(cur)
        cur["thresholds"] = [tuple(t) for t in cur["thresholds"]]
        out[cur["id"]] = cur
    if not out:
        # Fallback for repos that haven't run the migration: parse the legacy
        # workbooks in memory so the app stays usable.
        for key, xp in paths.CURRICULA.items():
            if os.path.exists(xp):
                out[key] = _legacy(key, xp)
    _cache["token"] = token
    _cache["data"] = out
    return out


def get(cid):
    return load_all().get(cid)


def save(cur):
    """Validate → backup the previous version → write. Returns the path."""
    validate(cur)
    stored = {k: v for k, v in cur.items() if k != "courses"}
    os.makedirs(paths.CURRICULA_DIR, exist_ok=True)
    path = _path(cur["id"])
    backups.backup(path)
    with open(path, "w", encoding="utf-8") as fh:
        json.dump(stored, fh, indent=2, ensure_ascii=False)
    _cache["token"] = None
    return path


def delete(cid):
    path = _path(cid)
    if not os.path.exists(path):
        raise CurriculumError(f"No curriculum '{cid}'.")
    backups.backup(path)
    os.remove(path)
    _cache["token"] = None


# ------------------------------------------------------------------- labels
def labels():
    """{id: display label}, e.g. {'2018': 'BSCS 2018–2024'}."""
    return {cid: label(cur) for cid, cur in load_all().items()}


# v1 header-table labels, kept so migration can map old MD fields to keys.
LEGACY_LABELS = {"2018": "2018–2024 (old)", "2025": "2025–present (new)"}


def key_from_label(text):
    """v1 label or key prefix -> curriculum key; None if unrecognized."""
    text = (text or "").strip()
    for key, lab in LEGACY_LABELS.items():
        if text == lab or text.startswith(key):
            return key
    for cid, lab in labels().items():
        if text == lab or text == cid:
            return cid
    return None


# ------------------------------------------------------------------- xlsx import
def parse_xlsx(path_or_stream):
    """Parse a Course-Checklist workbook -> (sections, thresholds).
    Layout: sheet 'Student', header on row 8, data from row 9; a 'Current
    Total Units' marker ends the course list, then (units, label) thresholds."""
    import openpyxl
    wb = openpyxl.load_workbook(path_or_stream, data_only=True)
    if "Student" not in wb.sheetnames:
        raise CurriculumError("Workbook has no 'Student' sheet — not a course checklist.")
    ws = wb["Student"]
    courses, thresholds = [], []
    year, cur_term = 1, None
    mode = "courses"
    for row in ws.iter_rows(min_row=9, values_only=True):
        a, b, c = row[0], row[1], row[2]
        units = row[5]
        if mode == "courses":
            if c is None or str(c).strip() == "":
                if row[4] and "Current Total Units" in str(row[4]):
                    mode = "between"
                continue
            t = str(a).strip() if a else ""
            if t and t != cur_term:
                if t == "1st" and cur_term in ("2nd", "Midyear"):
                    year += 1
                cur_term = t
            courses.append({
                "code": str(b).strip(), "title": re.sub(r"\s+", " ", str(c)).strip(),
                "units": float(units) if units not in (None, "") else 0.0,
                "prereq": re.sub(r"\s+", " ", str(row[6])).strip() if row[6] else "None",
                "year": year, "term": cur_term,
            })
        else:
            if a is not None and str(a).strip().replace(".", "").isdigit() and b:
                thresholds.append([float(a), str(b).strip()])
    if not courses:
        raise CurriculumError("No course rows found in the workbook.")
    sections, key = [], None
    for c in courses:
        if (c["year"], c["term"]) != key:
            key = (c["year"], c["term"])
            sections.append({"year": c["year"], "term": c["term"], "courses": []})
        sections[-1]["courses"].append({k: c[k] for k in ("code", "title", "units", "prereq")})
    return sections, sorted(thresholds)


def from_xlsx(path_or_stream, cid, program, start, end=None):
    from datetime import date
    sections, thresholds = parse_xlsx(path_or_stream)
    today = date.today().isoformat()
    return validate({
        "schema": SCHEMA, "id": cid, "program": program,
        "effective_start": start, "effective_end": end,
        "sections": sections, "thresholds": thresholds,
        "meta": {"created": today, "updated": today, "source": "xlsx import"},
    })


def _legacy(key, xlsx_path):
    sections, thresholds = parse_xlsx(xlsx_path)
    cur = {"schema": SCHEMA, "id": key, "program": "BSCS",
           "effective_start": int(key), "effective_end": 2024 if key == "2018" else None,
           "sections": sections, "thresholds": thresholds, "meta": {"source": "legacy xlsx"}}
    validate(cur)
    cur["courses"] = flatten(cur)
    cur["thresholds"] = [tuple(t) for t in cur["thresholds"]]
    return cur
